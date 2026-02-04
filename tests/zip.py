import zipfile
import os
import pytest
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
from io import BytesIO

files_path = os.path.join(os.getcwd(), "files")
pdf_path = os.path.join(files_path, "report.pdf")
xlsx_path = os.path.join(files_path, "B-4.xlsx")
csv_path = os.path.join(files_path, "QUOTA_FILE.csv")
zip_path = os.path.join(files_path, "zipfile.zip")

# Объявляем фикстуру
@pytest.fixture(scope="session")
def prepare_zip_archive():
    with zipfile.ZipFile(zip_path, 'w') as zf:
        for file in (pdf_path, xlsx_path, csv_path):
            zf.write(file, os.path.basename(file))
    yield
    os.remove(zip_path)

def test_check_pdf_file(prepare_zip_archive):
    with zipfile.ZipFile(zip_path, "r") as zp:
        with zp.open("report.pdf") as file:
            pdf_data = file.read()
            reader = PdfReader(BytesIO(pdf_data))

            # Проверка количества страниц
            number_of_pages = len(reader.pages)
            assert number_of_pages == 1, 'Ошибка в количестве страниц в файле Pdf'

            # Проверка содержания текста
            page = reader.pages[0]
            text = page.extract_text()
            assert """Пример документа в формате PDF""" in text, 'Ошибка в проверке содержимого PDF'


def test_check_csv_file(prepare_zip_archive):
    with zipfile.ZipFile(zip_path, "r") as zp:
        with zp.open("QUOTA_FILE.csv") as file:
            text = file.readline().decode('utf-8')
            assert "quotaCode" in text, 'Ошибка в тексте файла csv'


def test_check_xlsx_file(prepare_zip_archive):
    with (zipfile.ZipFile(zip_path, "r") as zp):
        with zp.open("B-4.xlsx") as file:
            text = load_workbook(file)
            sheet = text.active
            assert sheet.cell(row=7, column=2).value == 'только для зарегистрированных пользователей'

            headers = [cell.value for cell in sheet[1]]
            assert headers == ['Условия', 'Варианты условий', 1, 2, 3, 4, 5, 6, 7, 8, None],'Ошибка в заголовках столбцов'

